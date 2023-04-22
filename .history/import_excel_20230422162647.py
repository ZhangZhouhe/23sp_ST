import sqlite3
import openpyxl
lists=sqlite3.connect('data.db') 
c=lists.cursor()
listinsheet=openpyxl.load_workbook(r'C:\Users\VC\OneDrive\python2019\database\edit_42479_2019-01-14T152133.xlsx')
datainlist=listinsheet.active #获取excel文件当前表格
data_truck='''INSERT INTO mylist(list_id,list_node_id,doc_signature,doc_title,doc_category,doc_definition,doc_heading,doc_index,doc_col1,doc_col2,doc_col3,doc_col4,doc_col5,
doc_col6,doc_col7,doc_col8,doc_code) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''
for row in datainlist.iter_rows(min_row=2,max_col=17,max_row=datainlist.max_row): 
#使excel各行数据成为迭代器
    cargo=[cell.value for cell in row] #敲黑板！！使每行中单元格成为迭代器
    c.execute(data_truck,cargo) #敲黑板！写入一行数据到数据库中表mylist
lists.commit()
lists.close()
